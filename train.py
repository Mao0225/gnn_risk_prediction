import torch
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
import numpy as np
import os
import json
from openpyxl import Workbook, load_workbook


def train_model(exp_dir,model, train_data, val_data, config, save_dir):
    """
    训练模型并验证，同时保存最好和最后的模型，以及每一轮训练数据
    """
    optimizer = torch.optim.Adam(model.parameters(), lr=config['learning_rate'], weight_decay=1e-4)
    criterion = torch.nn.MSELoss()

    history = {
        "epoch": [],
        "train_loss": [],
        "val_loss": [],
        "val_MSE": [],
        "val_MAE": [],
        "val_R2": []
    }

    best_val_loss = float("inf")
    best_model_path = os.path.join(save_dir, "best_model.pth")
    last_model_path = os.path.join(save_dir, "last_model.pth")

    os.makedirs(save_dir, exist_ok=True)

    best_metrics = {}  # 记录最佳模型对应的指标

    for epoch in range(1, config['epochs'] + 1):
        # ---- Training ----
        model.train()
        optimizer.zero_grad()
        train_out = model(train_data)
        train_loss = criterion(train_out, train_data.y)
        train_loss.backward()
        optimizer.step()

        # ---- Validation ----
        model.eval()
        with torch.no_grad():
            val_out = model(val_data)
            val_loss = criterion(val_out, val_data.y)

            val_pred = val_out.detach().cpu().numpy()
            val_true = val_data.y.detach().cpu().numpy()

            mse = mean_squared_error(val_true, val_pred)
            mae = mean_absolute_error(val_true, val_pred)
            r2 = r2_score(val_true, val_pred)

        # ---- Save metrics ----
        history["epoch"].append(epoch)
        history["train_loss"].append(train_loss.item())
        history["val_loss"].append(val_loss.item())
        history["val_MSE"].append(mse)
        history["val_MAE"].append(mae)
        history["val_R2"].append(r2)

        if epoch % 50 == 0 or epoch == 1:
            print(f"Epoch {epoch}/{config['epochs']} | "
                  f"Train Loss: {train_loss.item():.4f} | "
                  f"Val Loss: {val_loss.item():.4f} | "
                  f"MSE: {mse:.4f} | MAE: {mae:.4f} | R²: {r2:.4f}")

        # ---- Save models ----
        torch.save(model.state_dict(), last_model_path)  # 每轮保存最后模型

        if val_loss.item() < best_val_loss:  # 如果验证集更好，保存最佳模型
            best_val_loss = val_loss.item()
            torch.save(model.state_dict(), best_model_path)
            best_metrics = {
                "val_loss": val_loss.item(),
                "MSE": mse,
                "MAE": mae,
                "R2": r2
            }

    # ---- Save full history ----
    history_path = os.path.join(save_dir, "training_history.json")
    with open(history_path, "w", encoding="utf-8") as f:
        json.dump(history, f, indent=4, ensure_ascii=False)

    # ---- Save best metrics & config into results.xlsx ----
    results_file = "results.xlsx"
    if not os.path.exists(results_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["exp_dir","hidden_channels", "learning_rate", "epochs",
                   "val_loss", "MSE", "MAE", "R2"])
        wb.save(results_file)

    wb = load_workbook(results_file)
    ws = wb["Results"]
    ws.append([
        exp_dir,
        config["hidden_channels"],
        config["learning_rate"],
        config["epochs"],
        best_metrics.get("val_loss", None),
        best_metrics.get("MSE", None),
        best_metrics.get("MAE", None),
        best_metrics.get("R2", None),
    ])
    wb.save(results_file)

    return history


def evaluate_model(model, data):
    """
    评估模型在给定数据上的性能
    """
    model.eval()
    with torch.no_grad():
        pred = model(data).detach().cpu().numpy()
        true = data.y.detach().cpu().numpy()

    mse = mean_squared_error(true, pred)
    mae = mean_absolute_error(true, pred)
    r2 = r2_score(true, pred)

    print(f'最终评估 - MSE: {mse:.4f}, MAE: {mae:.4f}, R²: {r2:.4f}')
    return pred


def save_model(model, path):
    """ 保存模型参数 """
    os.makedirs(os.path.dirname(path), exist_ok=True)
    torch.save(model.state_dict(), path)
